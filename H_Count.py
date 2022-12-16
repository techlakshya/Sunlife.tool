from tkinter import *
from tkinter import ttk
import openpyxl as op
import socket
from threading import Thread


class Traacker(Thread):
    
    def __init__(lak,root):
        Thread.__init__(lak)
        lak.root= root
        lak.root.title("Tracker")
        lak.root.geometry("510x255")
        lak.root.resizable(False,False)
        lak.Pcount1= StringVar()
        lak.Pcount2= StringVar()
        lak.Pcount3= StringVar()
        lak.Pcount4= StringVar()
        lak.Tcount1= StringVar()
        lak.Tcount2= StringVar()
        lak.Tcount3= StringVar()
        lak.Tcount4= StringVar()
        lak.login()

    def login(lak):
        lak.VM_num= socket.gethostname()
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        VM_no= S_UserDetail["C"]
        VM_list=[]
        for i in VM_no:
            VM_list.append(i.value)
        if lak.VM_num  in  VM_list:
            lak.track()
        else:
            Tracker_frame= Frame(lak.root,bg="#C352DA",relief= "raised",borderwidth=10)
            Tracker_frame.place(x=0,y=0,relwidth=1, relheight=1)
            Lable_tr=Label(lak.root,bg="black",relief= "raised",borderwidth=5,text="You are a new User.\nPlease ask your SME/Manager to update the record.",
            font=("Arial",15, "bold"),fg="white")
            Lable_tr.place(x=5,y=120)

    def track(lak):
        lak.Pcount1.set(0)
        lak.Pcount2.set(0)
        lak.Pcount3.set(0)
        lak.Pcount4.set(0)
        lak.Tcount1.set(0)
        lak.Tcount2.set(0)
        lak.Tcount3.set(0)
        lak.Tcount4.set(0)
        Tracker_frame= Frame(lak.root,bg="#461C4E",relief= "raised",borderwidth=5 )
        Tracker_frame.place(x=0,y=0,relwidth=1, relheight=1)
        user= socket.gethostname()
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        VM_no= S_UserDetail["C"]
        VM_list=[]
        for i in VM_no:
            VM_list.append(i.value)
        User= S_UserDetail["A"]
        User_list=[]
        for i in User:
            User_list.append(i.value)
        Team= S_UserDetail["D"]
        Team_list=[]
        for i in Team:
            Team_list.append(i.value)
        ACF2_no= S_UserDetail["B"]
        ACF2_list=[]
        for i in ACF2_no:
            ACF2_list.append(i.value)
        VM_ind=VM_list.index(user)
        lak.user= User_list[VM_ind]
        lak.Team= Team_list[VM_ind]
        lak.ACF2= ACF2_list[VM_ind]
        if lak.Team=="Payments & Taxation":
            A= "Payment-1st"
            B= "Payment-2nd"
            C= "Payment-3rd"
            D= "Payment-4th"
            E= "Taxation-1st"
            F= "Taxation-2nd"
            G= "Taxation-3rd"
            H= "Taxation-4th"
        else:
            A= "1st-Count"
            B= "2nd-Count"
            C= "3rd-Count"
            D= "4th-Count"
            E= "5th-Count"
            F= "6th-Count"
            G= "7th-Count"
            H= "8th-Count"
        
        Well_Lbl= Label(lak.root,text=("User :   {}      Team :   {}".format(lak.user,lak.Team)),justify="left" , bg="#461C4E", font = ("Arial",9, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=40,height=30)
        Well_Lbl= Label(lak.root,width=42,relief= "raised",text= "Count Report", bg="#461C4E", font = ("Arial",14, "bold"), fg= "white")
        Well_Lbl.place(x=2,y=2,height=40)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= A, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=71,height=35)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= B, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=106,height=35)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= C, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=141,height=35)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= D, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=176,height=35)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= E, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=260,y=71,height=35)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= F, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=260,y=106,height=35)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= G, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=260,y=141,height=35)
        Well_Lbl= Label(lak.root,width=11,relief= "raised",text= H, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=260,y=176,height=35)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Pcount1,  relief= "sunken", bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=130,y=71,height=32)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Pcount2,relief= "sunken",bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=130,y=106,height=32)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Pcount3,relief= "sunken", bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=130,y=141,height=32)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Pcount4,relief= "sunken", bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=130,y=176,height=32)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Tcount1,relief= "sunken", bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=380,y=71,height=32)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Tcount2,relief= "sunken",bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=380,y=106,height=32)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Tcount3,relief= "sunken", bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=380,y=141,height=32)
        Well_Lbl= Entry(lak.root,width=12,textvariable=lak.Tcount4,relief= "sunken", bg="white", font = ("Arial",13, "bold"), fg= "black",borderwidth=5)
        Well_Lbl.place(x=380,y=176,height=32)
        Well_Lbl= Button(lak.root,width=48,activebackground= "#7FB49A",activeforeground= "white",relief= "raised",
        text= "Update",command=lak.trackkerr, bg="#682F73", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=215)
        
    def trackkerr(lak):
        x= lak.ACF2
        x=x.upper()
        try:
            lak.NWB=op.load_workbook("Y:\ISS\BWMS\Q&N\KAT\{}_KAT.xlsx".format(x))
        except:
            lak.NWB=op.Workbook()
            lak.NWB.save("Y:\ISS\BWMS\Q&N\KAT\{}_KAT.xlsx".format(x))
        try:
            S_Notify= lak.NWB["Hourly Count"]
        except:
            lak.NWB.create_sheet("Hourly Count")
            S_Notify= lak.NWB["Hourly Count"]
            S_Notify["A1"]="ACF2"
            S_Notify["A2"]=x
            S_Notify["B1"]="Name"
            S_Notify["B2"]=lak.user
            S_Notify["C1"]="Count1"
            S_Notify["D1"]="Count2"
            S_Notify["E1"]="Count3"
            S_Notify["F1"]="Count4"
            S_Notify["G1"]="Count5"
            S_Notify["H1"]="Count6"
            S_Notify["I1"]="Count7"
            S_Notify["J1"]="Count8"
        S_Notify.cell(row=2, column=3).value= int(lak.Pcount1.get())
        S_Notify.cell(row=2, column=4).value= int(lak.Pcount2.get())
        S_Notify.cell(row=2, column=5).value= int(lak.Pcount3.get())
        S_Notify.cell(row=2, column=6).value= int(lak.Pcount4.get())
        S_Notify.cell(row=2, column=7).value= int(lak.Tcount1.get())
        S_Notify.cell(row=2, column=8).value= int(lak.Tcount2.get())
        S_Notify.cell(row=2, column=9).value= int(lak.Tcount3.get())
        S_Notify.cell(row=2, column=10).value= int(lak.Tcount4.get())

        lak.NWB.save("Y:\ISS\BWMS\Q&N\KAT\{}_KAT.xlsx".format(x))

if __name__== "__main__":
    root=Tk()
    aap= Traacker(root)
    aap.start()
    root.mainloop()