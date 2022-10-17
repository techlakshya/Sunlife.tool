import datetime
from tkinter import *
from tkinter import ttk
import tkinter
import openpyxl as op
import socket
from threading import Thread


class SMETOOL(Thread):
    
    def __init__(lak,root):
        Thread.__init__(lak)
        lak.root= root
        lak.root.title("SME_Tool")
        lak.root.geometry("1000x850")
        lak.root.resizable(False,False)
        lak.role= StringVar()
        lak.SME= StringVar()
        lak.WP= StringVar()
        lak.New_name= StringVar()
        lak.New_team= StringVar()
        lak.New_Afc2= StringVar()
        lak.Hidden= StringVar()
        lak.Hidden.set(0)
        lak.WPreply=StringVar()
        lak.replySME=StringVar()
        lak.updateNo= StringVar()
        lak.updateBestPract= StringVar()
        lak.login()

    def login(lak):
        lak.VM_num= socket.gethostname()
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        VM_no= S_UserDetail["C"]
        VM_list=[]
        for i in VM_no:
            VM_list.append(i.value)
        if lak.VM_num  in  VM_list:
            lak.main_screen()
        else:
            lak.login_screen()

    def login_screen(lak):
        MainFrame= Frame(lak.root ,borderwidth=0,bg= "black")
        MainFrame.place(x=0,y=0,relwidth=1, relheight=1)
        New_user= Label(MainFrame, text="New User" ,justify="left",bg= "black",fg="white",font = ("Comic Sans MS",22, "bold"),width=36 )
        New_user.place(x=228,y=310)
        New_user= Label(MainFrame,text= "Name:",justify="left",bg= "black" , font = ("Arial",14, "bold"), fg= "white")
        New_user.place(x=220,y=382)
        New_user= Label(MainFrame,text= "Team:",justify="left",bg= "black" , font = ("Arial",14, "bold"), fg= "white")
        New_user.place(x=220,y=462)
        New_user= Label(MainFrame,text="ACF2 ID:",justify="left",bg= "black" , font = ("Arial",14, "bold"), fg= "white")
        New_user.place(x=220,y=542)
        Count_entry =Entry(MainFrame ,textvariable=lak.New_name, width=35,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Count_entry.place(x=360,y=380)
        Count_entry =Entry(MainFrame ,textvariable=lak.New_team, width=35,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Count_entry.place(x=360,y=460)
        Count_entry =Entry(MainFrame ,textvariable=lak.New_Afc2, width=35,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Count_entry.place(x=360,y=540)
        Que_combo= ttk.Combobox(MainFrame,textvariable=lak.New_team, width=33,font = ("Arial",14, "bold"), state="readonly")
        Que_combo["values"]=("Payments & Taxation","Policy Titles", "Ins & Plan change & Illustration")
        Que_combo.set("Select")
        Que_combo.place(x=365,y=463)
        button4= Button(MainFrame,borderwidth=2,activebackground= "#7FB49A",activeforeground= "white", bg="black",
        text = " Submit ",command=lak.Login_submit, fg= "white",width=45,font = ("Arial", 14, "bold"))
        button4.place(x=220,y=618)

    def Login_submit(lak):
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        Final_list=[lak.New_name.get(), lak.New_Afc2.get(),lak.VM_num, lak.New_team.get()]
        S_UserDetail.append(Final_list)
        lak.UWB.save("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        lak.main_screen()

    def main_screen(lak):
        lak.Hidden.set(0)
        MainFrame= Frame(lak.root ,borderwidth=0, background="#120F7F")
        MainFrame.place(x=0,y=0,relwidth=1, relheight=1)
        user= socket.gethostname()
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        VM_no= S_UserDetail["C"]
        VM_list=[]
        for i in VM_no:
            VM_list.append(i.value)
        User= S_UserDetail["A"]
        User_list=[]
        for i in User:
            User_list.append(i.value)
        Team= S_UserDetail["D"]
        Team_list=[]
        for i in Team:
            Team_list.append(i.value)
        VM_ind=VM_list.index(user)
        lak.user= User_list[VM_ind]
        lak.Team= Team_list[VM_ind]
        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        S_Question=lak.QWB["Questions"]
        Csr_name=S_Question["C"]
        Csr_list=[]
        you_name_ind=[]
        for i in Csr_name:
            Csr_list.append(i.value)
        x=len(Csr_list)
        for i in range(x):
            if Csr_list[i]==lak.user:
                you_name_ind.append(i)
        ques=S_Question["F"]
        que_list=[]
        for i in ques:
            que_list.append(i.value)
        status=S_Question["G"]
        status_list=[]
        for i in status:
            status_list.append(i.value)
        reply=S_Question["H"]
        reply_list=[]
        for i in reply:
            reply_list.append(i.value)
        if lak.Team=="Payments & Taxation":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        #hidden:
        Well_Lbl= Label(lak.root,text= "User Name:",justify="left" , bg="#120F7F", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=820)
        Well_Lbl= Label(lak.root,text= lak.user,justify="left",bg="#120F7F",  font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=100,y=820)
        Well_Lbl= Label(lak.root,text= "Team Name:",justify="left",bg="#120F7F", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=250,y=820)
        Well_Lbl= Label(lak.root,text= lak.Team,justify="left",bg="#120F7F",  font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=350,y=820)
        Well_Lbl= Label(lak.root,relief= "sunken",text= "   Resource/Updates",width=58,borderwidth=6,justify="left", bg="#36338B", font = ("Arial",20, "bold"), fg= "white")
        Well_Lbl.place(x=0,y=0)
        Well_Lbl= Button(lak.root,relief= "raised",command=lak.pending, text= "Pending:",width=50,borderwidth=4,activebackground= "yellow",activeforeground= "#36338B",
        justify="left", bg="#36338B", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=500,y=48)
        Well_Lbl= Button(lak.root,relief= "raised",command=lak.Closed, text= "All Questions:",width=50,borderwidth=4,activebackground= "yellow",activeforeground= "#36338B",
        justify="left", bg="#36338B", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=5,y=48)
        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=0,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=996,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left",bg="black"  )
        Well_Lbl.place(x=0,y=805,width=1100,height=2)
        button4= Button(lak.root,borderwidth=4,activebackground= "yellow",activeforeground= "#36338B",bg="#120F7F",
        text = " Approves ",command=lak.Approval,fg= "white",width=26,font = ("Arial", 12, "bold"))
        button4.place(x=723,y=810)
        lak.pending()
        
    def pending(lak):

        Pend1Fram= Frame(lak.root, borderwidth=5,bg= "#120F7F")
        Pend1Fram.place(x=0, y=85, width=1000,height=560)
        PendFram= Frame(lak.root, borderwidth=5,bg= "#504D9F")
        PendFram.place(x=0, y=240, width=1000,height=560) 
        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        S_Question=lak.QWB["Questions"]
        Csr_name=S_Question["C"]
        Csr_list=[]
        you_name_ind=[]
        for i in Csr_name:
            Csr_list.append(i.value)
        SME_name=S_Question["J"]
        SME_list=[]
        you_name_ind=[]
        for i in SME_name:
            SME_list.append(i.value)
        x=len(SME_list)
        for i in range(x):
            if SME_list[i]==lak.user:
                you_name_ind.append(i)
        ques=S_Question["F"]
        que_list=[]
        for i in ques:
            que_list.append(i.value)
        status=S_Question["G"]
        status_list=[]
        for i in status:
            status_list.append(i.value)
        reply=S_Question["H"]
        reply_list=[]
        for i in reply:
            reply_list.append(i.value)
        Wp_list=[]
        Wp_num= S_Question["E"]
        for i in Wp_num:
            Wp_list.append(i.value)    

        pending_index_list=[]
        for i in range(x):
            if status_list[i]=="Pending":
                pending_index_list.append(i)
        pno=len(pending_index_list)
        Well_Lbl= Label(lak.root,text=pno,justify="left", bg="#36338B", font = ("Arial",13, "bold"), fg= "white")
        Well_Lbl.place(x=800,y=55)
        #pending index----------------------------
        try:
            a=pending_index_list[0]
        except:
            pass

        try:
            c=pending_index_list[1]
        except:
            pass
        try:
            d=pending_index_list[2]
        except:
            pass
        try:
            e=pending_index_list[3]
        except:
            pass
        try:
            f=pending_index_list[4]
        except:
            pass
        try:
            g=pending_index_list[5]
        except:
            pass
        try:
            h=pending_index_list[6]
        except:
            pass
        try:
            j=pending_index_list[7]
        except:
            pass
        try:
            k=pending_index_list[8]
        except:
            pass
        try:
            l=pending_index_list[9]
        except:
            pass


        try:
            CSRName1= Csr_list[a]
            Workpackage1=Wp_list[a]
            question1=que_list[a]
            solution1=reply_list[a]
        except:
            pass
        try:
            CSRName2= Csr_list[c]
            Workpackage2=Wp_list[c]
            question2=que_list[c]
            solution2=reply_list[c]
        except:
            pass
        try:
            CSRName3= Csr_list[d]
            Workpackage3=Wp_list[d]
            question3= que_list[d]
            solution3=reply_list[d]
        except:
            pass
        try:
            CSRName4= Csr_list[e]
            Workpackage4=Wp_list[e]
            question4=que_list[e]
            solution4=reply_list[e]
        except:
            pass
        try:
            CSRName5= Csr_list[f]
            Workpackage5=Wp_list[f]
            question5=que_list[f]
            solution5=reply_list[f]
        except:
            pass
        try:
            CSRName6= Csr_list[g]
            Workpackage6=Wp_list[g]
            question6=que_list[g]
            solution6=reply_list[g]
        except:
            pass
        try:
            CSRName7= Csr_list[h]
            Workpackage7=Wp_list[h]
            question7=que_list[h]
            solution7=reply_list[h]
        except:
            pass
        try:
            CSRName8= Csr_list[j]
            Workpackage8=Wp_list[j]
            question8=que_list[j]
            solution8=reply_list[j]
        except:
            pass
        try:
            CSRName9= Csr_list[k]
            Workpackage9=Wp_list[k]
            question9=que_list[k]
            solution9=reply_list[k]
        except:
            pass
        #-------------------------------------------

        Well_Lbl= Label(Pend1Fram,relief= "raised",text= "Work_Package",font = ("Arial",12, "bold"),width=20, fg= "white",bg="#093375")
        Well_Lbl.place(x=4,y=0,height=40)
        Well_Lbl= Button(Pend1Fram,relief= "raised",command=lak.replyupdate, borderwidth=6, text= "Reply",font = ("Arial",11, "bold"),width=21, fg= "white",bg="#093375")
        Well_Lbl.place(x=4,y=45,height=100)
        Well_Lbl= Entry(Pend1Fram, relief= "sunken",textvariable=lak.WP, borderwidth=5, font = ("Arial", 11, "bold"),width=92, fg= "black")
        Well_Lbl.place(x=230,y=0,height=40)
        lak.Text_box= Text(Pend1Fram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=92, fg= "black")
        lak.Text_box.place(x=230,y=45,height=100)
        
        Well_Lbl= Label(PendFram,relief= "raised",text= "CSR Name",font = ("Arial",12, "bold"),width=16, fg= "white",bg="#093375")
        Well_Lbl.place(x=2,y=0,height=40)
        Well_Lbl= Label(PendFram,relief= "raised",text= "Questions",font = ("Arial",12, "bold"),width=41, fg= "white",bg="#093375")
        Well_Lbl.place(x=167,y=0,height=40)
        Well_Lbl= Label(PendFram,relief= "raised",text= "Solutions",font = ("Arial",12, "bold"),width=41, fg= "white",bg="#093375")
        Well_Lbl.place(x=580,y=0,height=40)
        try:
            Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= CSRName1,font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=46,height=150)
            Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
            Well_Lbl.place(x=167,y=46,height=150)
            Well_Lbl.insert(tkinter.INSERT, Workpackage1+"\n")
            Well_Lbl.insert(tkinter.INSERT,"\n"+question1)
            Well_Lbl.config(state="disabled")
            Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
            Well_Lbl.place(x=580,y=46,height=150)
            Well_Lbl.insert(tkinter.INSERT,"\n"+solution1)
        except:
            pass
        try:   
            Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= CSRName2,font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=200,height=150)
            Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
            Well_Lbl.place(x=167,y=200,height=150)
            Well_Lbl.insert(tkinter.INSERT, Workpackage2 +"\n")
            Well_Lbl.insert(tkinter.INSERT,"\n"+question2)
            Well_Lbl.config(state="disabled")
            Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
            Well_Lbl.place(x=580,y=200,height=150)
            Well_Lbl.insert(tkinter.INSERT,"\n"+solution2)
        except:
            pass
        try:    
            Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= CSRName3,font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=355,height=150)
            Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
            Well_Lbl.place(x=167,y=355,height=150)
            Well_Lbl.insert(tkinter.INSERT, Workpackage3+"\n")
            Well_Lbl.insert(tkinter.INSERT,"\n"+question3)
            Well_Lbl.config(state="disabled")
            Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
            Well_Lbl.place(x=580,y=355,height=150)
            Well_Lbl.insert(tkinter.INSERT,"\n"+solution3)
        except:
            pass
        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=0,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=996,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left",bg="black"  )
        Well_Lbl.place(x=0,y=805,width=1100,height=2)
        def Pending2():
            PendFram= Frame(lak.root, borderwidth=5,bg= "#504D9F")
            PendFram.place(x=0, y=240, width=1000,height=520) 
            Well_Lbl= Label(PendFram,relief= "raised",text= "CSR Name",font = ("Arial",12, "bold"),width=16, fg= "white",bg="#093375")
            Well_Lbl.place(x=2,y=0,height=40)
            Well_Lbl= Label(PendFram,relief= "raised",text= "Questions",font = ("Arial",12, "bold"),width=41, fg= "white",bg="#093375")
            Well_Lbl.place(x=167,y=0,height=40)
            Well_Lbl= Label(PendFram,relief= "raised",text= "Solutions",font = ("Arial",12, "bold"),width=41, fg= "white",bg="#093375")
            Well_Lbl.place(x=580,y=0,height=40)

            try: 
                Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
                text= CSRName4,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=2,y=46,height=150)
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
                Well_Lbl.place(x=167,y=46,height=150)
                Well_Lbl.insert(tkinter.INSERT, Workpackage4+"\n")
                Well_Lbl.insert(tkinter.INSERT,"\n"+question4)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
                Well_Lbl.place(x=580,y=46,height=150)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution4)
            except:
                pass
            try: 
                Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
                text= CSRName5,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=2,y=200,height=150)
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
                Well_Lbl.place(x=167,y=200,height=150)
                Well_Lbl.insert(tkinter.INSERT, Workpackage5 +"\n")
                Well_Lbl.insert(tkinter.INSERT,"\n"+question5)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
                Well_Lbl.place(x=580,y=200,height=150)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution5)
            except:
                pass
            try:
                Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
                text= CSRName6,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=2,y=355,height=150)
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
                Well_Lbl.place(x=167,y=355,height=150)
                Well_Lbl.insert(tkinter.INSERT, Workpackage6+"\n")
                Well_Lbl.insert(tkinter.INSERT,"\n"+question6)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
                Well_Lbl.place(x=580,y=355,height=150)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution6)
            except:
                pass
            Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
            Well_Lbl.place(x=0,y=0,height=1000)
            Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
            Well_Lbl.place(x=996,y=0,height=1000)
            Well_Lbl= Label(lak.root,justify="left",bg="black"  )
            Well_Lbl.place(x=0,y=805,width=1100,height=2)

        def Pending3():
            PendFram= Frame(lak.root, borderwidth=5,bg= "#504D9F")
            PendFram.place(x=0, y=240, width=1000,height=520) 
            Well_Lbl= Label(PendFram,relief= "raised",text= "CSR Name",font = ("Arial",12, "bold"),width=16, fg= "white",bg="#093375")
            Well_Lbl.place(x=2,y=0,height=40)
            Well_Lbl= Label(PendFram,relief= "raised",text= "Questions",font = ("Arial",12, "bold"),width=41, fg= "white",bg="#093375")
            Well_Lbl.place(x=167,y=0,height=40)
            Well_Lbl= Label(PendFram,relief= "raised",text= "Solutions",font = ("Arial",12, "bold"),width=41, fg= "white",bg="#093375")
            Well_Lbl.place(x=580,y=0,height=40)
            
            try:
                Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
                text= CSRName7,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=2,y=46,height=150)
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
                Well_Lbl.place(x=167,y=46,height=150)
                Well_Lbl.insert(tkinter.INSERT, Workpackage7+"\n")
                Well_Lbl.insert(tkinter.INSERT,"\n"+question7)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
                Well_Lbl.place(x=580,y=46,height=150)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution7)
            except:
                pass
            try:
                Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
                text= CSRName8,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=2,y=200,height=150)
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
                Well_Lbl.place(x=167,y=200,height=150)
                Well_Lbl.insert(tkinter.INSERT, Workpackage8 +"\n")
                Well_Lbl.insert(tkinter.INSERT,"\n"+question8)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
                Well_Lbl.place(x=580,y=200,height=150)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution8)
            except:
                pass
            try:
                Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
                text= CSRName9,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=2,y=355,height=150)
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=50, fg= "black")
                Well_Lbl.place(x=167,y=355,height=150)
                Well_Lbl.insert(tkinter.INSERT, Workpackage9+"\n")
                Well_Lbl.insert(tkinter.INSERT,"\n"+question9)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(PendFram,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=50, fg= "green")
                Well_Lbl.place(x=580,y=355,height=150)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution9)
            except:
                pass
            Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
            Well_Lbl.place(x=0,y=0,height=1000)
            Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
            Well_Lbl.place(x=996,y=0,height=1000)
            Well_Lbl= Label(lak.root,justify="left",bg="black"  )
            Well_Lbl.place(x=0,y=805,width=1100,height=2)

        Well_Lbl= Button(PendFram,relief= "raised",command=lak.pending, text= "Page1:",width=33,activebackground= "yellow",activeforeground= "#36338B",
        borderwidth=4,justify="left", bg="#36338B", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=0,y=514)
        Well_Lbl= Button(PendFram,relief= "raised",command=Pending2, text= "Page2:",width=33,activebackground= "yellow",activeforeground= "#36338B",
        borderwidth=4,justify="left", bg="#36338B", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=330,y=514)
        Well_Lbl= Button(PendFram,relief= "raised",command=Pending3, text= "Page3:",width=33,activebackground= "yellow",activeforeground= "#36338B",
        borderwidth=4,justify="left", bg="#36338B", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=660,y=514)

        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=0,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=996,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left",bg="black"  )
        Well_Lbl.place(x=0,y=805,width=1100,height=2)

    def replyupdate(lak):

        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        S_Question=lak.QWB["Questions"]
        Status_list=[]
        Status= S_Question["G"]
        for i in  Status:
            Status_list.append(i.value)
        WP_list=[]
        WP_rep= S_Question["E"]
        for i in  WP_rep:
            WP_list.append(i.value)
        SME_reply=[]
        WP_rep= S_Question["H"]
        for i in  WP_rep:
            SME_reply.append(i.value)
        Closingtime=[]
        WP_rep= S_Question["I"]
        for i in  WP_rep:
            Closingtime.append(i.value)
        ClosingDate=[]
        WP_rep= S_Question["K"]
        for i in  WP_rep:
            ClosingDate.append(i.value)
        x= datetime.datetime.now()
        Date=x.strftime("%d/%m/%Y")
        hour=str(x.strftime("%H"))
        min=str(x.minute)
        Raised_time= hour+":"+min
        Q= WP_list.index(lak.WP.get())
        R= Q+1
        S_Question.cell(row=R, column=7).value= "Closed"
        S_Question.cell(row=R, column=8).value= lak.Text_box.get(1.0,END)
        S_Question.cell(row=R, column=9).value= Raised_time
        S_Question.cell(row=R, column=11).value=Date

        if lak.Team=="Payments & Taxation":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        lak.pending()
        lak.WP.set("")

    def Closed(lak):
        sub_fram= Frame(master=lak.root, bg="#194D33")
        sub_fram.place(x=0, y=90, width=1000,height=715) 
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Date",font = ("Arial",12, "bold"),width=10, fg= "white")
        Well_Lbl.place(x=5,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "CSR Name",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=111,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Question",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=408,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Solution",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=705,y=5,height=33)
        #-------------------------------------------------------------------------------------------------------------------------------------------------
        Z= int(lak.Hidden.get())
        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        S_Notify=lak.QWB["Questions"]
        Serial=S_Notify["C"]
        CSR_list=[]
        for i in Serial:
            CSR_list.append(i.value)
        WP_list=S_Notify["E"]
        R_list=[]    #list of WP
        CSR_ind=[]
        for i in WP_list:
            R_list.append(i.value)
        for i in R_list:
            CSR_ind.append(R_list.index(i))
        Date=S_Notify["A"]
        S_list=[]  #list of dated
        for i in Date:
            S_list.append(i.value)
        Solution=S_Notify["F"]
        E_list=[]  #list of question
        for i in Solution:
            E_list.append(i.value)
        ans=S_Notify["H"]
        F_list=[]  #list of solution
        for i in ans:
            F_list.append(i.value)


        try:
            SNEntry1=CSR_list[Z+1]
            REntry1= R_list[Z+1]
            SEntry1=S_list[Z+1]
            EEntery1= E_list[Z+1]
            FEntry1=F_list[Z+1]
            CSR_ind1=CSR_ind[Z+1]
        except:
            pass
        try:
            SNEntry2=CSR_list[Z+2]
            REntry2= R_list[Z+2]
            SEntry2=S_list[Z+2]
            EEntery2= E_list[Z+2]
            FEntry2=F_list[Z+2]
            CSR_ind2=CSR_ind[Z+2]
        except:
            pass
        try:
            SNEntry3=CSR_list[Z+3]
            REntry3= R_list[Z+3]
            SEntry3=S_list[Z+3]
            EEntery3= E_list[Z+3]
            FEntry3=F_list[Z+3]
            CSR_ind3=CSR_ind[Z+3]
        except:
            pass
        try:
            SNEntry4=CSR_list[Z+4]
            REntry4= R_list[Z+4]
            SEntry4=S_list[Z+4]
            EEntery4= E_list[Z+4]
            FEntry4=F_list[Z+4]
            CSR_ind4=CSR_ind[Z+4]
        except:
            pass
        try:            
            SNEntry5=CSR_list[Z+5]
            REntry5= R_list[Z+5]
            SEntry5=S_list[Z+5]
            EEntery5= E_list[Z+5]
            FEntry5=F_list[Z+5]
            CSR_ind5=CSR_ind[Z+5]
        except:
            pass
        
        lak.Hidden.set(Z+5)
            #--------------------------------------------------------------------------------------------------------------------------------------------------
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry1,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=42, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=42 ,height=120)
            Text1.insert(tkinter.INSERT,REntry1+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry1)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=42, height=120)
            Text2.insert(tkinter.INSERT,EEntery1)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=42, height=120)
            Text3.insert(tkinter.INSERT,FEntry1)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry2,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=167, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=167 ,height=120)
            Text1.insert(tkinter.INSERT,REntry2+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry2)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=167, height=120)
            Text2.insert(tkinter.INSERT,EEntery2)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=167, height=120)
            Text3.insert(tkinter.INSERT,FEntry2)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry3,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=292, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=292 ,height=120)
            Text1.insert(tkinter.INSERT,REntry3+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry3)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=292, height=120)
            Text2.insert(tkinter.INSERT,EEntery3)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=292, height=120)
            Text3.insert(tkinter.INSERT,FEntry3)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry4,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=417, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=417 ,height=120)
            Text1.insert(tkinter.INSERT,REntry4+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry4)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=417, height=120)
            Text2.insert(tkinter.INSERT,EEntery4)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=417, height=120)
            Text3.insert(tkinter.INSERT,FEntry4)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5, bg="#A1D89E",text=SEntry5,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=542, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=542 ,height=120)
            Text1.insert(tkinter.INSERT,REntry5+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry5)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=542, height=120)
            Text2.insert(tkinter.INSERT,EEntery5)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=542, height=120)
            Text3.insert(tkinter.INSERT,FEntry5)
            Text3.config(state="disable")        
        except:
            pass
        
        Role_entry =Entry(sub_fram , width=5,textvariable=lak.Hidden ,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Role_entry.place(x=925,y=5)
        Well_Lbl= Label(sub_fram,justify="left" ,bg="#194D33" )
        Well_Lbl.place(x=995,y=0,height=1000)
        Well_Lbl= Label(sub_fram,justify="left" ,bg="#194D33" )
        Well_Lbl.place(x=995,y=0,height=1000)
        button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
        text = " Next ",command=lak.Notify_Elements,fg= "white", bg="#194D33",width=82,font = ("Arial", 15, "bold"))
        button4.place(x=5,y=755)
        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=0,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
        Well_Lbl.place(x=996,y=0,height=1000) 

    def Notify_Elements(lak):   
        sub_fram= Frame(master=lak.root, bg="#194D33")
        sub_fram.place(x=0, y=90, width=1000,height=715)          
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Date",font = ("Arial",12, "bold"),width=10, fg= "white")
        Well_Lbl.place(x=5,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "CSR Name",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=111,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Question",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=408,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Solution",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=705,y=5,height=33)
        #-------------------------------------------------------------------------------------------------------------------------------------------------
        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        S_Notify=lak.QWB["Questions"]
        Serial=S_Notify["C"]
        CSR_list=[]
        for i in Serial:
            CSR_list.append(i.value)
        WP_list=S_Notify["E"]
        R_list=[]    #list of WP
        CSR_ind=[]
        for i in WP_list:
            R_list.append(i.value)
        for i in R_list:
            CSR_ind.append(R_list.index(i))
        Date=S_Notify["A"]
        S_list=[]  #list of dated
        for i in Date:
            S_list.append(i.value)
        Solution=S_Notify["F"]
        E_list=[]  #list of question
        for i in Solution:
            E_list.append(i.value)
        ans=S_Notify["H"]
        F_list=[]  #list of solution
        for i in ans:
            F_list.append(i.value)
        Z= int(lak.Hidden.get())
        
        try:
            SNEntry1=CSR_list[Z+1]
            REntry1= R_list[Z+1]
            SEntry1=S_list[Z+1]
            EEntery1= E_list[Z+1]
            FEntry1=F_list[Z+1]
            CSR_ind1=CSR_ind[Z+1]
        except:
            pass
        try:
            SNEntry2=CSR_list[Z+2]
            REntry2= R_list[Z+2]
            SEntry2=S_list[Z+2]
            EEntery2= E_list[Z+2]
            FEntry2=F_list[Z+2]
            CSR_ind2=CSR_ind[Z+2]
        except:
            pass
        try:
            SNEntry3=CSR_list[Z+3]
            REntry3= R_list[Z+3]
            SEntry3=S_list[Z+3]
            EEntery3= E_list[Z+3]
            FEntry3=F_list[Z+3]
            CSR_ind3=CSR_ind[Z+3]
        except:
            pass
        try:
            SNEntry4=CSR_list[Z+4]
            REntry4= R_list[Z+4]
            SEntry4=S_list[Z+4]
            EEntery4= E_list[Z+4]
            FEntry4=F_list[Z+4]
            CSR_ind4=CSR_ind[Z+4]
        except:
            pass
        try:            
            SNEntry5=CSR_list[Z+5]
            REntry5= R_list[Z+5]
            SEntry5=S_list[Z+5]
            EEntery5= E_list[Z+5]
            FEntry5=F_list[Z+5]
            CSR_ind5=CSR_ind[Z+5]
        except:
            pass
        Z=lak.Hidden.set(Z+5)
        
            #--------------------------------------------------------------------------------------------------------------------------------------------------
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry1,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=42, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=42 ,height=120)
            Text1.insert(tkinter.INSERT,REntry1+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry1)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=42, height=120)
            Text2.insert(tkinter.INSERT,EEntery1)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=42, height=120)
            Text3.insert(tkinter.INSERT,FEntry1)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry2,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=167, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=167 ,height=120)
            Text1.insert(tkinter.INSERT,REntry2+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry2)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=167, height=120)
            Text2.insert(tkinter.INSERT,EEntery2)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=167, height=120)
            Text3.insert(tkinter.INSERT,FEntry2)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry3,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=292, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=292 ,height=120)
            Text1.insert(tkinter.INSERT,REntry3+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry3)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=292, height=120)
            Text2.insert(tkinter.INSERT,EEntery3)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=292, height=120)
            Text3.insert(tkinter.INSERT,FEntry3)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text=SEntry4,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=417, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=417 ,height=120)
            Text1.insert(tkinter.INSERT,REntry4+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry4)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=417, height=120)
            Text2.insert(tkinter.INSERT,EEntery4)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=417, height=120)
            Text3.insert(tkinter.INSERT,FEntry4)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5, bg="#A1D89E",text=SEntry5,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=542, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=542 ,height=120)
            Text1.insert(tkinter.INSERT,REntry5+":\n\n")
            Text1.insert(tkinter.INSERT,SNEntry5)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=542, height=120)
            Text2.insert(tkinter.INSERT,EEntery5)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=542, height=120)
            Text3.insert(tkinter.INSERT,FEntry5)
            Text3.config(state="disable")        
        except:
            pass
        
        Role_entry =Entry(sub_fram , width=5,textvariable=lak.Hidden ,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Role_entry.place(x=925,y=5)
        Well_Lbl= Label(sub_fram,justify="left" ,bg="#194D33" )
        Well_Lbl.place(x=995,y=0,height=1000)
        button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
        text = " Next ",command=lak.Closed,fg= "white", bg="#194D33",width=82,font = ("Arial", 15, "bold"))
        button4.place(x=5,y=755)


    def Approval(lak):
            PendFram= Frame(root, borderwidth=5,bg= "#504D9F")
            PendFram.place(x=0, y=90, width=1000,height=715) 
            #Role	Situation	English Notification	French  Notification
            Well_Lbl= Label(lak.root,relief= "sunken",text= "Notification:",width=58,borderwidth=6,justify="left", bg="#142167", font = ("Arial",20, "bold"), fg= "white")
            Well_Lbl.place(x=0,y=95)
            Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "Role",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=50,height=40)
            Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "Situation",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=95,height=70)
            Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "English",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=170,height=70)
            Well_Lbl= Label(PendFram, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "French",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=245,height=70)
            Entrybox= Entry(PendFram,textvariable=lak.updateNo, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            Entrybox.place(x=170,y=50,height=40)
            lak.Text2App= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            lak.Text2App.place(x=170,y=95,height=70)
            lak.Text3App= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            lak.Text3App.place(x=170,y=170,height=70)
            lak.Text4App= Text(PendFram,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            lak.Text4App.place(x=170,y=245,height=70)
            #S.No	Role	Situation	Confusion	Solution	Please enter 1
            Well_Lbl= Label(lak.root,relief= "sunken",text= "Best Practice:",width=58,borderwidth=6, bg="#142167", font = ("Arial",20, "bold"), fg= "white")
            Well_Lbl.place(x=0,y=450)
            Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "Role",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=500,height=40)
            Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "Situation",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=545,height=70)
            Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "Confusion",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=620,height=70)
            Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",borderwidth=5,activebackground= "yellow",activeforeground= "#36338B",
            text= "Solution",font = ("Arial",12, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=2,y=695,height=70)
            Entrybox= Entry(lak.root,textvariable=lak.updateBestPract, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            Entrybox.place(x=170,y=500,height=40)
            lak.Text5App= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            lak.Text5App.place(x=170,y=545,height=70)
            lak.Text6App= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            lak.Text6App.place(x=170,y=620,height=70)
            lak.Text7App= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=101, fg= "black")
            lak.Text7App.place(x=170,y=695,height=70)          
            button4= Button(lak.root,borderwidth=4,activebackground= "yellow",activeforeground= "#36338B",bg="#120F7F",
            text = " Update ",command=lak.Notifyupdate,fg= "white",width=100,font = ("Arial", 12, "bold"))
            button4.place(x=2,y=412)
            button4= Button(lak.root,borderwidth=4,activebackground= "yellow",activeforeground= "#36338B",bg="#120F7F",
            text = " Update ",command=lak.BestPracUpdate,fg= "white",width=100,font = ("Arial", 12, "bold"))
            button4.place(x=2,y=768)
            Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
            Well_Lbl.place(x=0,y=0,height=1000)
            Well_Lbl= Label(lak.root,justify="left" ,bg="black" )
            Well_Lbl.place(x=996,y=0,height=1000) 
            
    def Notifyupdate(lak):
        lak.NWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\Notification_DB.xlsx")
        if lak.Team=="Payments & Taxation":
            S_Notify= lak.NWB["Payments & Taxation"]
        elif lak.Team=="Policy Titles":
            S_Notify= lak.NWB["Titles"]
        elif lak.Team=="Ins & Plan change & Illustration":
            S_Notify= lak.NWB["Ins & Plan change &Illustration"]
        noti= S_Notify["B"]
        noti_list=[]
        for i in  noti:
            noti_list.append(i.value)
        Q= noti_list.index("Blank")+1

        S_Notify.cell(row=Q, column=2).value= lak.updateNo.get()
        S_Notify.cell(row=Q, column=3).value= lak.Text2App.get(1.0,END)
        S_Notify.cell(row=Q, column=4).value= lak.Text3App.get(1.0,END)
        S_Notify.cell(row=Q, column=5).value= lak.Text4App.get(1.0,END)
        S_Notify.cell(row=Q, column=6).value= lak.user
        Q=Q+1
        for i in range(11):
            S_Notify.cell(row=Q, column=2).value= "Blank"
            S_Notify.cell(row=Q, column=3).value= "Blank"
            S_Notify.cell(row=Q, column=4).value= "Blank"
            S_Notify.cell(row=Q, column=5).value= "Blank"
            S_Notify.cell(row=Q, column=6).value= "Blank"
            Q= (Q+1)
        lak.NWB.save("Y:\\ISS\\BWMS\\Q&N\\Notification_DB.xlsx")
        lak.Approval()
        lak.updateNo.set("")

    def BestPracUpdate(lak):
        lak.BPWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\BestPrac_DB.xlsx")
        if lak.Team=="Payments & Taxation":
            S_Notify= lak.BPWB["Payments & Taxation"]
        elif lak.Team=="Policy Titles":
            S_Notify= lak.BPWB["Titles"]
        elif lak.Team=="Ins & Plan change & Illustration":
            S_Notify= lak.BPWB["Ins & Plan change &Illustration"]
        noti= S_Notify["B"]
        noti_list=[]
        for i in  noti:
            noti_list.append(i.value)
        Q= noti_list.index("Blank")+1
        S_Notify.cell(row=Q, column=2).value= lak.updateBestPract.get()
        S_Notify.cell(row=Q, column=3).value= lak.Text5App.get(1.0,END)
        S_Notify.cell(row=Q, column=4).value= lak.Text6App.get(1.0,END)
        S_Notify.cell(row=Q, column=5).value= lak.Text7App.get(1.0,END)
        S_Notify.cell(row=Q, column=6).value= 1
        S_Notify.cell(row=Q, column=7).value= lak.user
        Q=Q+1
        for i in range(11):
            S_Notify.cell(row=Q, column=2).value= "Blank"
            S_Notify.cell(row=Q, column=3).value= "Blank"
            S_Notify.cell(row=Q, column=4).value= "Blank"
            S_Notify.cell(row=Q, column=5).value= "Blank"
            S_Notify.cell(row=Q, column=6).value= "Blank"
            S_Notify.cell(row=Q, column=7).value= "Blank"
            Q= (Q+1)
        lak.BPWB.save("Y:\\ISS\\BWMS\\Q&N\\BestPrac_DB.xlsx")
        lak.Approval()
        lak.updateBestPract.set("")



if __name__== "__main__":
    root=Tk()
    aap= SMETOOL(root)
    aap.start()
    root.mainloop()