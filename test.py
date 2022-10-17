import datetime
from tkinter import *
from tkinter import ttk
import tkinter
import openpyxl as op
import socket
from threading import Thread
import random

class Notification(Thread):
    
    def __init__(lak,root):
        Thread.__init__(lak)
        lak.root= root
        lak.root.title("Handbook")
        lak.root.geometry("1000x850")
        lak.root.resizable(False,False)
        lak.role= StringVar()
        lak.SME= StringVar()
        lak.WP= StringVar()
        lak.New_name= StringVar()
        lak.New_team= StringVar()
        lak.New_Afc2= StringVar()
        lak.Hidden= StringVar()
        lak.Hidden.set(0)
        lak.login()

        # lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        # lak.NWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\Notification_DB.xlsx")
        # lak.BPWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\BestPrac_DB.xlsx")

        # if lak.Team=="Payments & Taxation":
        #     lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        # elif lak.Team=="Policy Titles":
        #     lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        # elif lak.Team=="Ins & Plan change & Illustration":
        #     lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        #Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx

    def login(lak):
        lak.VM_num= socket.gethostname()
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        VM_no= S_UserDetail["C"]
        VM_list=[]
        for i in VM_no:
            VM_list.append(i.value)
        if lak.VM_num  in  VM_list:
            lak.main_screen()
        else:
            lak.login_screen()

    def login_screen(lak):
        MainFrame= Frame(lak.root ,borderwidth=0,bg= "black")
        MainFrame.place(x=0,y=0,relwidth=1, relheight=1)
        New_user= Label(MainFrame, text="New User" ,justify="left",bg= "black",fg="white",font = ("Comic Sans MS",22, "bold"),width=36 )
        New_user.place(x=228,y=310)
        New_user= Label(MainFrame,text= "Name:",justify="left",bg= "black" , font = ("Arial",14, "bold"), fg= "white")
        New_user.place(x=220,y=382)
        New_user= Label(MainFrame,text= "Team:",justify="left",bg= "black" , font = ("Arial",14, "bold"), fg= "white")
        New_user.place(x=220,y=462)
        New_user= Label(MainFrame,text="ACF2 ID:",justify="left",bg= "black" , font = ("Arial",14, "bold"), fg= "white")
        New_user.place(x=220,y=542)
        Count_entry =Entry(MainFrame ,textvariable=lak.New_name, width=35,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Count_entry.place(x=360,y=380)
        Count_entry =Entry(MainFrame ,textvariable=lak.New_team, width=35,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Count_entry.place(x=360,y=460)
        Count_entry =Entry(MainFrame ,textvariable=lak.New_Afc2, width=35,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
        Count_entry.place(x=360,y=540)
        Que_combo= ttk.Combobox(MainFrame,textvariable=lak.New_team, width=33,font = ("Arial",14, "bold"), state="readonly")
        Que_combo["values"]=("Payments & Taxation","Policy Titles", "Ins & Plan change & Illustration")
        Que_combo.set("Select")
        Que_combo.place(x=365,y=463)
        button4= Button(MainFrame,borderwidth=2,activebackground= "#7FB49A",activeforeground= "white", bg="black",
        text = " Submit ",command=lak.Login_submit, fg= "white",width=45,font = ("Arial", 14, "bold"))
        button4.place(x=220,y=618)

    def Login_submit(lak):
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        Final_list=[lak.New_name.get(), lak.New_Afc2.get(),lak.VM_num, lak.New_team.get()]
        S_UserDetail.append(Final_list)
        lak.UWB.save("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        lak.main_screen()

    def main_screen(lak):
        lak.Hidden.set(0)
        MainFrame= Frame(lak.root ,borderwidth=0)
        MainFrame.place(x=0,y=0,relwidth=1, relheight=1)
        user= socket.gethostname()
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        VM_no= S_UserDetail["C"]
        VM_list=[]
        for i in VM_no:
            VM_list.append(i.value)
        User= S_UserDetail["A"]
        User_list=[]
        for i in User:
            User_list.append(i.value)
        Team= S_UserDetail["D"]
        Team_list=[]
        for i in Team:
            Team_list.append(i.value)
        VM_ind=VM_list.index(user)
        lak.user= User_list[VM_ind]
        lak.Team= Team_list[VM_ind]
        SMEs= S_UserDetail["E"]
        sl=[]
        for i in SMEs:
            sl.append(i.value)
        SME= tuple(sl)
        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")        

        S_Question=lak.QWB["Questions"]
        Csr_name=S_Question["C"]
        Csr_list=[]
        you_name_ind=[]
        for i in Csr_name:
            Csr_list.append(i.value)
        x=len(Csr_list)
        for i in range(x):
            if Csr_list[i]==lak.user:
                you_name_ind.append(i)
        ques=S_Question["F"]
        que_list=[]
        for i in ques:
            que_list.append(i.value)
        status=S_Question["G"]
        status_list=[]
        for i in status:
            status_list.append(i.value)
        reply=S_Question["H"]
        reply_list=[]
        for i in reply:
            reply_list.append(i.value)
        if lak.Team=="Payments & Taxation":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")
        #hidden:
        Well_Lbl= Label(lak.root,justify="left" ,bg="#56635C" )
        Well_Lbl.place(x=667,y=3,height=280)
        Well_Lbl= Label(lak.root,justify="left" ,bg="#56635C" )
        Well_Lbl.place(x=0,y=0,height=1000)
        Well_Lbl= Label(lak.root,justify="left" ,bg="#56635C")
        Well_Lbl.place(x=0,y=318,width=1100,height=4)
        Well_Lbl= Label(lak.root,justify="left",bg="#69716D"  )
        Well_Lbl.place(x=0,y=805,width=1100,height=2)
        Well_Lbl= Label(lak.root,justify="left" ,bg="#69716D" )
        Well_Lbl.place(x=0,y=280,width=1100,height=4)
        Well_Lbl= Label(lak.root,justify="left" ,bg="#69716D" )
        Well_Lbl.place(x=0,y=37,width=1100,height=4)
        COI_entry =Entry(lak.root , width=50,borderwidth=7,foreground= "black",font = ("Arial", 12, "bold"))
        COI_entry.place(x=200,y=50)
        COI_entry =Entry(lak.root , width=50,borderwidth=7,foreground= "black",font = ("Arial", 12, "bold"))
        COI_entry.place(x=200,y=90)
        Well_Lbl= Label(lak.root,text= "User Name:",justify="left" , font = ("Arial",12, "bold"), fg= "black")
        Well_Lbl.place(x=10,y=820)
        Well_Lbl= Label(lak.root,text= lak.user,justify="left", font = ("Arial",12, "bold"), fg= "black")
        Well_Lbl.place(x=100,y=820)
        Well_Lbl= Label(lak.root,text= "Team Name:",justify="left",font = ("Arial",12, "bold"), fg= "black")
        Well_Lbl.place(x=250,y=820)
        Well_Lbl= Label(lak.root,text= lak.Team,justify="left", font = ("Arial",12, "bold"), fg= "black")
        Well_Lbl.place(x=350,y=820)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Ask Question:",width=55,borderwidth=6,justify="left", font = ("Arial",15, "bold"), fg= "black")
        Well_Lbl.place(x=0,y=0)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Health Tips:",width=26,borderwidth=6,justify="left", font = ("Arial",15, "bold"), fg= "black")
        Well_Lbl.place(x=673,y=0)
        Well_Lbl= Button(lak.root,command=lak.YourQuestion ,text= "Your Questions:",borderwidth=4,justify="left", font = ("Arial",13,"bold"), fg= "black")
        Well_Lbl.place(x=5,y=283)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Role:",font = ("Arial",12, "bold"),width=15, fg= "black")
        Well_Lbl.place(x=10,y=58)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Send To:", font = ("Arial",12, "bold"),width=15, fg= "black")
        Well_Lbl.place(x=10,y=95)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Work_Package:", font = ("Arial",12, "bold"),width=15, fg= "black")
        Well_Lbl.place(x=10,y=128)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Question:", font = ("Arial",12, "bold"),width=15, fg= "black")
        Well_Lbl.place(x=10,y=164)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Date",font = ("Arial",12, "bold"),width=15, fg= "black")
        Well_Lbl.place(x=5,y=322)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Status",font = ("Arial",12, "bold"),width=15, fg= "black")
        Well_Lbl.place(x=160,y=322)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Questions",font = ("Arial",12, "bold"),width=34, fg= "black")
        Well_Lbl.place(x=316,y=322)
        Well_Lbl= Label(lak.root,relief= "raised",text= "Solutions",font = ("Arial",12, "bold"),width=33, fg= "black")
        Well_Lbl.place(x=662,y=322)
        Que_combo= ttk.Combobox(lak.root,textvariable=lak.role, width=54,font = ("Arial",11, "bold") )
        Que_combo["values"]=("",)
        Que_combo.set("")
        Que_combo.place(x=205,y=54)
        Que_combo= ttk.Combobox(lak.root,textvariable=lak.SME, width=54,font = ("Arial",11, "bold") )
        Que_combo["values"]=SME
        Que_combo.set("Resource Person")
        Que_combo.place(x=205,y=95)
        COI_entry =Entry(lak.root ,textvariable=lak.WP, width=50,borderwidth=7,foreground= "black",font = ("Arial", 12, "bold"))
        COI_entry.place(x=200,y=125)
        lak.Text_box =Text(lak.root, wrap=WORD, width=50,borderwidth=7 ,foreground= "black",font = ("Arial", 12))
        lak.Text_box.place(x=200,y=160, height=120)
        button4= Button(lak.root ,activebackground= "#7FB49A",activeforeground= "white",
        text = "Submit",command=lak.Submit, fg= "black",width=15,font = ("Arial", 13, "bold"))
        button4.place(x=10,y=198, height=78)
        button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
        text = " Notification ",command=lak.Notify,fg= "black",width=26,font = ("Arial", 12, "bold"))
        button4.place(x=723,y=810)
        #Random Best Practices
        lak.BPWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\BestPrac_DB.xlsx")
        if lak.Team=="Payments & Taxation":
            S_Notify= lak.BPWB["Payments & Taxation"]
        elif lak.Team=="Policy Titles":
            S_Notify= lak.BPWB["Titles"]
        elif lak.Team=="Ins & Plan change & Illustration":
            S_Notify= lak.BPWB["Ins & Plan change &Illustration"]
        sitution= S_Notify["C"]
        Sit_list=[]
        for i in sitution:
            Sit_list.append(i.value)
        Bestprac= S_Notify["E"]  
        Bestprac_list=[]
        for i in Bestprac:
            Bestprac_list.append(i.value) 
        ran= S_Notify["F"]
        Ran_list=[]
        for i in ran:
            if i.value==1:
                Ran_list.append(i.value) 
        ind=len(Ran_list)
        l= random.randint(1,ind)
        c1= ("Situatuion:")
        t1=("\n"+Sit_list[l]+"\n\n")
        c2="What need to do?\n"
        t2=Bestprac_list[l]
        Text1 =Text(lak.root , wrap=WORD, width=35,borderwidth=5,foreground= "blue",font = ("Arial", 12,"bold"))
        Text1.place(x=670,y=40 ,height=200)
        Text1.insert(tkinter.INSERT,c1)
        Text1.insert(tkinter.INSERT,t1)
        Text1.insert(tkinter.INSERT,c2)
        Text1.insert(tkinter.INSERT,t2)
        Text1.config(state="disable")
        button4= Button(lak.root,borderwidth=4, activebackground= "#7FB49A",activeforeground= "white",
        text = " Explore More ",command=lak.Best_Pract,fg= "black",width=31,font = ("Arial", 12, "bold"))
        button4.place(x=673,y=242)
        Well_Lbl= Label(lak.root,justify="left" ,bg="#56635C" )
        Well_Lbl.place(x=995,y=0,height=1000)
    #______________________________________________________________________________________________second Screen______________________________________________________________________________
    def Notify(lak):
        MainFrame= Frame(lak.root ,borderwidth=0, bg="black")
        MainFrame.place(x=0,y=0,relwidth=1, relheight=1)
        Noti_labl= Label(master= lak.root,relief="raise", text="Notification",bg="#181B19", fg="white", font=("Arial",18, "bold") ,borderwidth=5, width=66)
        Noti_labl.place(x=0,y=5)
        sub_fram= Frame(master=lak.root, bg="#194D33")
        sub_fram.place(x=0,y=82, relheight=1, relwidth=1)
        Well_Lbl= Label(lak.root,justify="left",bg= "black")
        Well_Lbl.place(x=0,y=80,width=1500,height=5)
        Well_Lbl= Label(lak.root, bg="#194D33",text= "User Name:",justify="left" , font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=820)
        Well_Lbl= Label(lak.root, bg="#194D33",text= lak.user,justify="left", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=100,y=820)
        Well_Lbl= Label(lak.root, bg="#194D33",text= "Team Name:",justify="left",font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=250,y=820)
        Well_Lbl= Label(lak.root, bg="#194D33",text= lak.Team,justify="left", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=350,y=820)
        Well_Lbl= Label(lak.root,justify="left",bg="#69716D"  )
        Well_Lbl.place(x=0,y=805,width=1100,height=2)
        button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
        text = " Ask Me ",command=lak.main_screen,fg= "white", bg="#194D33",width=26,font = ("Arial", 12, "bold"))
        button4.place(x=723,y=810)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "S. No.",font = ("Arial",12, "bold"),width=10, fg= "white")
        Well_Lbl.place(x=5,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Situations",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=111,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "In English",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=408,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "In French",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=705,y=5,height=33)
        #-------------------------------------------------------------------------------------------------------------------------------------------------
        Z= int(lak.Hidden.get())
        lak.NWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\Notification_DB.xlsx")
        if lak.Team=="Payments & Taxation":
            S_Notify= lak.NWB["Payments & Taxation"]
        elif lak.Team=="Policy Titles":
            S_Notify= lak.NWB["Titles"]
        elif lak.Team=="Ins & Plan change & Illustration":
            S_Notify= lak.NWB["Ins & Plan change &Illustration"]
        Serial=S_Notify["A"]
        SN_list=[]
        for i in Serial:
            SN_list.append(i.value)
        Role=S_Notify["B"]
        R_list=[]
        for i in Role:
            R_list.append(i.value)
        Situation=S_Notify["C"]
        S_list=[]
        for i in Situation:
            S_list.append(i.value)
        E_Notification=S_Notify["D"]
        E_list=[]
        for i in E_Notification:
            E_list.append(i.value)
        F_Notification=S_Notify["E"]
        F_list=[]
        for i in F_Notification:
            F_list.append(i.value)

        try:
            SNEntry1=SN_list[Z+1]
            REntry1= R_list[Z+1]
            SEntry1=S_list[Z+1]
            EEntery1= E_list[Z+1]
            FEntry1=F_list[Z+1]
        except:
            pass
        try:
            SNEntry2=SN_list[Z+2]
            REntry2= R_list[Z+2]
            SEntry2=S_list[Z+2]
            EEntery2= E_list[Z+2]
            FEntry2=F_list[Z+2]
        except:
            pass
        try:
            SNEntry3=SN_list[Z+3]
            REntry3= R_list[Z+3]
            SEntry3=S_list[Z+3]
            EEntery3= E_list[Z+3]
            FEntry3=F_list[Z+3]
        except:
            pass
        try:
            SNEntry4=SN_list[Z+4]
            REntry4= R_list[Z+4]
            SEntry4=S_list[Z+4]
            EEntery4= E_list[Z+4]
            FEntry4=F_list[Z+4]
        except:
            pass
        try:
            SNEntry5=SN_list[Z+5]
            REntry5= R_list[Z+5]
            SEntry5=S_list[Z+5]
            EEntry5= E_list[Z+5]
            FEntry5=F_list[Z+5]
        except:
            pass
            #--------------------------------------------------------------------------------------------------------------------------------------------------
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry1,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=42, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=42 ,height=120)
            Text1.insert(tkinter.INSERT,REntry1+":\n\n")
            Text1.insert(tkinter.INSERT,SEntry1)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=42, height=120)
            Text2.insert(tkinter.INSERT,EEntery1)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=42, height=120)
            Text3.insert(tkinter.INSERT,FEntry1)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry2,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=167, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=167 ,height=120)
            Text1.insert(tkinter.INSERT,REntry2+":\n\n")
            Text1.insert(tkinter.INSERT,SEntry2)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=167, height=120)
            Text2.insert(tkinter.INSERT,EEntery2)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=167, height=120)
            Text3.insert(tkinter.INSERT,FEntry2)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry3,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=292, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=292 ,height=120)
            Text1.insert(tkinter.INSERT,REntry3+":\n\n")
            Text1.insert(tkinter.INSERT,SEntry3)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=292, height=120)
            Text2.insert(tkinter.INSERT,EEntery3)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=292, height=120)
            Text3.insert(tkinter.INSERT,FEntry3)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry4,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=417, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=417 ,height=120)
            Text1.insert(tkinter.INSERT,REntry4+":\n\n")
            Text1.insert(tkinter.INSERT,SEntry4)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=417, height=120)
            Text2.insert(tkinter.INSERT,EEntery4)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=417, height=120)
            Text3.insert(tkinter.INSERT,FEntry4)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5, bg="#A1D89E",text= SNEntry5,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=542, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=542 ,height=120)
            Text1.insert(tkinter.INSERT,REntry5+":\n\n")
            Text1.insert(tkinter.INSERT,SEntry5)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=542, height=120)
            Text2.insert(tkinter.INSERT,EEntry5)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=542, height=120)
            Text3.insert(tkinter.INSERT,FEntry5)
            Text3.config(state="disable")        
        except:
            pass

        Role_entry =Entry(lak.root , width=5,textvariable=lak.Hidden ,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"), state="readonly")
        Role_entry.place(x=925,y=48)
        lak.Hidden.set(SNEntry5)
        Well_Lbl= Label(sub_fram,justify="left" ,bg="#194D33" )
        Well_Lbl.place(x=995,y=0,height=1000)
        Z= int(lak.Hidden.get())
        def Notify_Elements():            
            sub_fram= Frame(master=lak.root, bg="#194D33")
            sub_fram.place(x=0,y=82, relwidth=1, relheight=1)
            Well_Lbl= Label(lak.root,justify="left",bg= "black")
            Well_Lbl.place(x=0,y=80,width=1500,height=5)
            lak.NWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\Notification_DB.xlsx")
            if lak.Team=="Payments & Taxation":
                S_Notify= lak.NWB["Payments & Taxation"]
            elif lak.Team=="Policy Titles":
                S_Notify= lak.NWB["Titles"]
            elif lak.Team=="Ins & Plan change & Illustration":
                S_Notify= lak.NWB["Ins & Plan change &Illustration"]
            Serial=S_Notify["A"]
            SN_list=[]
            for i in Serial:
                SN_list.append(i.value)
            Role=S_Notify["B"]
            R_list=[]
            for i in Role:
                R_list.append(i.value)
            Situation=S_Notify["C"]
            S_list=[]
            for i in Situation:
                S_list.append(i.value)
            E_Notification=S_Notify["D"]
            E_list=[]
            for i in E_Notification:
                E_list.append(i.value)
            F_Notification=S_Notify["E"]
            F_list=[]
            for i in F_Notification:
                F_list.append(i.value)
            try:
                SNEntry1=SN_list[Z+1]
                REntry1= R_list[Z+1]
                SEntry1=S_list[Z+1]
                EEntery1= E_list[Z+1]
                FEntry1=F_list[Z+1]
            except:
                pass
            try:
                SNEntry2=SN_list[Z+2]
                REntry2= R_list[Z+2]
                SEntry2=S_list[Z+2]
                EEntery2= E_list[Z+2]
                FEntry2=F_list[Z+2]
            except:
                pass
            try:
                SNEntry3=SN_list[Z+3]
                REntry3= R_list[Z+3]
                SEntry3=S_list[Z+3]
                EEntery3= E_list[Z+3]
                FEntry3=F_list[Z+3]
            except:
                pass
            try:
                SNEntry4=SN_list[Z+4]
                REntry4= R_list[Z+4]
                SEntry4=S_list[Z+4]
                EEntery4= E_list[Z+4]
                FEntry4=F_list[Z+4]
            except:
                pass
            try:
                SNEntry5=SN_list[Z+5]
                REntry5= R_list[Z+5]
                SEntry5=S_list[Z+5]
                EEntery5= E_list[Z+5]
                FEntry5=F_list[Z+5]
            except:
                pass
            lak.Hidden.set(SNEntry5)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "S. No.",font = ("Arial",12, "bold"),width=10, fg= "white")
            Well_Lbl.place(x=5,y=5,height=33)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Situations",font = ("Arial",12, "bold"),width=29, fg= "white")
            Well_Lbl.place(x=111,y=5,height=33)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "In English",font = ("Arial",12, "bold"),width=29, fg= "white")
            Well_Lbl.place(x=408,y=5,height=33)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "In French",font = ("Arial",12, "bold"),width=29, fg= "white")
            Well_Lbl.place(x=705,y=5,height=33)
            Role_entry =Entry(lak.root , width=5,textvariable=lak.Hidden ,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
            Role_entry.place(x=925,y=48)
            lak.Hidden.set(SNEntry5)
            Well_Lbl= Label(lak.root, bg="#194D33",text= "User Name:",justify="left" , font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=10,y=820)
            Well_Lbl= Label(lak.root, bg="#194D33",text= lak.user,justify="left", font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=100,y=820)
            Well_Lbl= Label(lak.root, bg="#194D33",text= "Team Name:",justify="left",font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=250,y=820)
            Well_Lbl= Label(lak.root, bg="#194D33",text= lak.Team,justify="left", font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=350,y=820)

            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry1,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=42, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=42 ,height=120)
                Text1.insert(tkinter.INSERT,REntry1+":\n\n")
                Text1.insert(tkinter.INSERT,SEntry1)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=42, height=120)
                Text2.insert(tkinter.INSERT,EEntery1)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=42, height=120)
                Text3.insert(tkinter.INSERT,FEntry1)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry2,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=167, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=167 ,height=120)
                Text1.insert(tkinter.INSERT,REntry2+":\n\n")
                Text1.insert(tkinter.INSERT,SEntry2)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=167, height=120)
                Text2.insert(tkinter.INSERT,EEntery2)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=167, height=120)
                Text3.insert(tkinter.INSERT,FEntry2)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry3,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=292, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=292 ,height=120)
                Text1.insert(tkinter.INSERT,REntry3+":\n\n")
                Text1.insert(tkinter.INSERT,SEntry3)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=292, height=120)
                Text2.insert(tkinter.INSERT,EEntery3)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=292, height=120)
                Text3.insert(tkinter.INSERT,FEntry3)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#A1D89E",text= SNEntry4,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=417, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=417 ,height=120)
                Text1.insert(tkinter.INSERT,SEntry4)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=417, height=120)
                Text1.insert(tkinter.INSERT,REntry4+":\n\n")
                Text2.insert(tkinter.INSERT,EEntery4)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=417, height=120)
                Text3.insert(tkinter.INSERT,FEntry4)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5, bg="#A1D89E",text= SNEntry5,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=542, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=542 ,height=120)
                Text1.insert(tkinter.INSERT,REntry5+":\n\n")
                Text1.insert(tkinter.INSERT,SEntry5)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=542, height=120)
                Text2.insert(tkinter.INSERT,EEntery5)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#A1D89E",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=542, height=120)
                Text3.insert(tkinter.INSERT,FEntry5)
                Text3.config(state="disable")  
            except:
                pass
            
            Well_Lbl= Label(sub_fram,justify="left" ,bg="#194D33" )
            Well_Lbl.place(x=995,y=0,height=1000)
            button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
            text = " Next ",command=lak.Notify,fg= "white", bg="#194D33",width=82,font = ("Arial", 15, "bold"))
            button4.place(x=5,y=755)
            Well_Lbl= Label(sub_fram,justify="left" ,bg="#194D33" )
            Well_Lbl.place(x=0,y=805,width=1100,height=2)
            button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
            text = " Ask Me ",command=lak.main_screen,fg= "white", bg="#194D33",width=26,font = ("Arial", 12, "bold"))
            button4.place(x=723,y=810)
            Well_Lbl= Label(lak.root,justify="left",bg="#69716D"  )
            Well_Lbl.place(x=0,y=805,width=1100,height=2)
                
        Well_Lbl= Label(sub_fram,justify="left" ,bg="#194D33" )
        Well_Lbl.place(x=995,y=0,height=1000)
        button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
        text = " Next ",command=Notify_Elements,fg= "white", bg="#194D33",width=82,font = ("Arial", 15, "bold"))
        button4.place(x=5,y=755)
        #______________________________________________________________________________________________Thired Screen_____________________________________________________________________________
    def Best_Pract(lak):
        MainFrame= Frame(lak.root ,borderwidth=0, bg="black")
        MainFrame.place(x=0,y=0,relwidth=1, relheight=1)
        Noti_labl= Label(master= lak.root,relief="raise", text="Best Practices",bg="#181B19", fg="white", font=("Arial",18, "bold") ,borderwidth=5, width=66)
        Noti_labl.place(x=0,y=5)
        sub_fram= Frame(master=lak.root, bg="#5D1C5C")
        sub_fram.place(x=0,y=82, relheight=1, relwidth=1)
        Well_Lbl= Label(lak.root,justify="left",bg= "black")
        Well_Lbl.place(x=0,y=80,width=1500,height=5)
        Well_Lbl= Label(lak.root, bg="#5D1C5C",text= "User Name:",justify="left" , font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=10,y=820)
        Well_Lbl= Label(lak.root, bg="#5D1C5C",text= lak.user,justify="left", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=100,y=820)
        Well_Lbl= Label(lak.root, bg="#5D1C5C",text= "Team Name:",justify="left",font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=250,y=820)
        Well_Lbl= Label(lak.root, bg="#5D1C5C",text= lak.Team,justify="left", font = ("Arial",12, "bold"), fg= "white")
        Well_Lbl.place(x=350,y=820)
        Well_Lbl= Label(lak.root,justify="left",bg="#69716D"  )
        Well_Lbl.place(x=0,y=805,width=1100,height=2)
        button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
        text = " Ask Me ",command=lak.main_screen,fg= "white", bg="#5D1C5C",width=26,font = ("Arial", 12, "bold"))
        button4.place(x=723,y=810)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Role",font = ("Arial",12, "bold"),width=10, fg= "white")
        Well_Lbl.place(x=5,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Situations",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=111,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Confusion",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=408,y=5,height=33)
        Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Best Practice",font = ("Arial",12, "bold"),width=29, fg= "white")
        Well_Lbl.place(x=705,y=5,height=33)
        #-------------------------------------------------------------------------------------------------------------------------------------------------
        Z= int(lak.Hidden.get())
        lak.BPWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\BestPrac_DB.xlsx")
        if lak.Team=="Payments & Taxation":
            S_Notify= lak.BPWB["Payments & Taxation"]
        elif lak.Team=="Policy Titles":
            S_Notify= lak.BPWB["Titles"]
        elif lak.Team=="Ins & Plan change & Illustration":
            S_Notify= lak.BPWB["Ins & Plan change &Illustration"]

        Serial=S_Notify["A"]
        SN_list=[]
        for i in Serial:
            SN_list.append(i.value)
        Role=S_Notify["B"]
        R_list=[]
        for i in Role:
            R_list.append(i.value)
        Situation=S_Notify["C"]
        S_list=[]
        for i in Situation:
            S_list.append(i.value)
        E_Notification=S_Notify["D"]
        E_list=[]
        for i in E_Notification:
            E_list.append(i.value)
        F_Notification=S_Notify["E"]
        F_list=[]
        for i in F_Notification:
            F_list.append(i.value)
        try:
            SNEntry1=SN_list[Z+1]
            REntry1= R_list[Z+1]
            SEntry1=S_list[Z+1]
            EEntery1= E_list[Z+1]
            FEntry1=F_list[Z+1]
        except:
            pass
        try:
            SNEntry2=SN_list[Z+2]
            REntry2= R_list[Z+2]
            SEntry2=S_list[Z+2]
            EEntery2= E_list[Z+2]
            FEntry2=F_list[Z+2]
        except:
            pass
        try:
            SNEntry3=SN_list[Z+3]
            REntry3= R_list[Z+3]
            SEntry3=S_list[Z+3]
            EEntery3= E_list[Z+3]
            FEntry3=F_list[Z+3]
        except:
            pass
        try:
            SNEntry4=SN_list[Z+4]
            REntry4= R_list[Z+4]
            SEntry4=S_list[Z+4]
            EEntery4= E_list[Z+4]
            FEntry4=F_list[Z+4]
        except:
            pass
        try:
            SNEntry5=SN_list[Z+5]
            REntry5= R_list[Z+5]
            SEntry5=S_list[Z+5]
            EEntry5= E_list[Z+5]
            FEntry5=F_list[Z+5]
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text=SNEntry1 ,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=42, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=42 ,height=120)
            Text1.insert(tkinter.INSERT,REntry1+"\n\n")
            Text1.insert(tkinter.INSERT,SEntry1)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=42, height=120)
            Text2.insert(tkinter.INSERT,EEntery1)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=42, height=120)
            Text3.insert(tkinter.INSERT,FEntry1)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text= SNEntry2,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=167, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=167 ,height=120)
            Text1.insert(tkinter.INSERT,REntry2+"\n\n")
            Text1.insert(tkinter.INSERT,SEntry2)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=167, height=120)
            Text2.insert(tkinter.INSERT,EEntery2)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=167, height=120)
            Text3.insert(tkinter.INSERT,FEntry2)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text= SNEntry3,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=292, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=292 ,height=120)
            Text1.insert(tkinter.INSERT,REntry3+"\n\n")
            Text1.insert(tkinter.INSERT,SEntry3)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=292, height=120)
            Text2.insert(tkinter.INSERT,EEntery3)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=292, height=120)
            Text3.insert(tkinter.INSERT,FEntry3)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text= SNEntry4,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=417, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=417 ,height=120)
            Text1.insert(tkinter.INSERT,REntry4+"\n\n")
            Text1.insert(tkinter.INSERT,SEntry4)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=417, height=120)
            Text2.insert(tkinter.INSERT,EEntery4)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=417, height=120)
            Text3.insert(tkinter.INSERT,FEntry4)
            Text3.config(state="disable")
        except:
            pass
        try:
            Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5, bg="#C595C5",text= SNEntry5,font = ("Arial",12, "bold"),width=10, fg= "black")
            Well_Lbl.place(x=5,y=542, height=120)
            Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text1.place(x=111,y=542 ,height=120)
            Text1.insert(tkinter.INSERT,REntry5+"\n\n")
            Text1.insert(tkinter.INSERT,SEntry5)
            Text1.config(state="disable")
            Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text2.place(x=408,y=542, height=120)
            Text2.insert(tkinter.INSERT,EEntry5)
            Text2.config(state="disable")
            Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
            Text3.place(x=705,y=542, height=120)
            Text3.insert(tkinter.INSERT,FEntry5)
            Text3.config(state="disable")        
        except:
            pass
        
        #hidden---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        Role_entry =Entry(lak.root ,textvariable=lak.Hidden, width=5 ,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"), state="readonly")
        Role_entry.place(x=925,y=48)
        lak.Hidden.set(SNEntry5)
        Well_Lbl= Label(sub_fram,justify="left" ,bg="#5D1C5C" )
        Well_Lbl.place(x=995,y=0,height=1000)
        Z= int(lak.Hidden.get())
        def Best_Pract_Element():            
            sub_fram= Frame(master=lak.root, bg="#5D1C5C")
            sub_fram.place(x=0,y=82, relwidth=1, relheight=1)
            Well_Lbl= Label(lak.root,justify="left",bg= "black")
            Well_Lbl.place(x=0,y=80,width=1500,height=5)
            lak.BPWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\BestPrac_DB.xlsx")
            if lak.Team=="Payments & Taxation":
                S_Notify= lak.BPWB["Payments & Taxation"]
            elif lak.Team=="Policy Titles":
                S_Notify= lak.BPWB["Titles"]
            elif lak.Team=="Ins & Plan change & Illustration":
                S_Notify= lak.BPWB["Ins & Plan change &Illustration"]

            Serial=S_Notify["A"]
            SN_list=[]
            for i in Serial:
                SN_list.append(i.value)
            Role=S_Notify["B"]
            R_list=[]
            for i in Role:
                R_list.append(i.value)
            Situation=S_Notify["C"]
            S_list=[]
            for i in Situation:
                S_list.append(i.value)
            E_Notification=S_Notify["D"]
            E_list=[]
            for i in E_Notification:
                E_list.append(i.value)
            F_Notification=S_Notify["E"]
            F_list=[]
            for i in F_Notification:
                F_list.append(i.value)
            try:
                SNEntry1=SN_list[Z+1]
                REntry1= R_list[Z+1]
                SEntry1=S_list[Z+1]
                EEntery1= E_list[Z+1]
                FEntry1=F_list[Z+1]
            except:
                pass
            try:
                SNEntry2=SN_list[Z+2]
                REntry2= R_list[Z+2]
                SEntry2=S_list[Z+2]
                EEntery2= E_list[Z+2]
                FEntry2=F_list[Z+2]
            except:
                pass
            try:
                SNEntry3=SN_list[Z+3]
                REntry3= R_list[Z+3]
                SEntry3=S_list[Z+3]
                EEntery3= E_list[Z+3]
                FEntry3=F_list[Z+3]
            except:
                pass
            try:
                SNEntry4=SN_list[Z+4]
                REntry4= R_list[Z+4]
                SEntry4=S_list[Z+4]
                EEntery4= E_list[Z+4]
                FEntry4=F_list[Z+4]
            except:
                pass
            try:
                SNEntry5=SN_list[Z+5]
                REntry5= R_list[Z+5]
                SEntry5=S_list[Z+5]
                EEntery5= E_list[Z+5]
                FEntry5=F_list[Z+5]
            except:
                pass

            lak.Hidden.set(SNEntry5)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Role",font = ("Arial",12, "bold"),width=10, fg= "white")
            Well_Lbl.place(x=5,y=5,height=33)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Situations",font = ("Arial",12, "bold"),width=29, fg= "white")
            Well_Lbl.place(x=111,y=5,height=33)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Confusion",font = ("Arial",12, "bold"),width=29, fg= "white")
            Well_Lbl.place(x=408,y=5,height=33)
            Well_Lbl= Label(sub_fram,relief= "raised",bg="black",text= "Best Practice",font = ("Arial",12, "bold"),width=29, fg= "white")
            Well_Lbl.place(x=705,y=5,height=33)
            Role_entry =Entry(lak.root , width=5,textvariable=lak.Hidden ,borderwidth=5,foreground= "black",font = ("Arial", 14, "bold"))
            Role_entry.place(x=925,y=48)
            lak.Hidden.set(SNEntry5)
            Well_Lbl= Label(lak.root, bg="#5D1C5C",text= "User Name:",justify="left" , font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=10,y=820)
            Well_Lbl= Label(lak.root, bg="#5D1C5C",text= lak.user,justify="left", font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=100,y=820)
            Well_Lbl= Label(lak.root, bg="#5D1C5C",text= "Team Name:",justify="left",font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=250,y=820)
            Well_Lbl= Label(lak.root, bg="#5D1C5C",text= lak.Team,justify="left", font = ("Arial",12, "bold"), fg= "white")
            Well_Lbl.place(x=350,y=820)
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text= SNEntry1,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=42, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=42 ,height=120)
                Text1.insert(tkinter.INSERT,REntry1+"\n\n")
                Text1.insert(tkinter.INSERT,SEntry1)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=42, height=120)
                Text2.insert(tkinter.INSERT,EEntery1)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=42, height=120)
                Text3.insert(tkinter.INSERT,FEntry1)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text= SNEntry2,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=167, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=167 ,height=120)
                Text1.insert(tkinter.INSERT,REntry2+"\n\n")
                Text1.insert(tkinter.INSERT,SEntry2)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=167, height=120)
                Text2.insert(tkinter.INSERT,EEntery2)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=167, height=120)
                Text3.insert(tkinter.INSERT,FEntry2)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text= SNEntry3,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=292, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=292 ,height=120)
                Text1.insert(tkinter.INSERT,REntry3+"\n\n")
                Text1.insert(tkinter.INSERT,SEntry3)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=292, height=120)
                Text2.insert(tkinter.INSERT,EEntery3)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=292, height=120)
                Text3.insert(tkinter.INSERT,FEntry3)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5,bg="#C595C5",text= SNEntry4,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=417, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=417 ,height=120)
                Text1.insert(tkinter.INSERT,REntry4+"\n\n")
                Text1.insert(tkinter.INSERT,SEntry4)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=417, height=120)
                Text2.insert(tkinter.INSERT,EEntery4)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=417, height=120)
                Text3.insert(tkinter.INSERT,FEntry4)
                Text3.config(state="disable")
            except:
                pass
            try:
                Well_Lbl= Label(sub_fram,relief= "sunken",borderwidth=5, bg="#C595C5",text= SNEntry5,font = ("Arial",12, "bold"),width=10, fg= "black")
                Well_Lbl.place(x=5,y=542, height=120)
                Text1 =Text(sub_fram , wrap=WORD, width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text1.place(x=111,y=542 ,height=120)
                Text1.insert(tkinter.INSERT,SEntry5)
                Text1.config(state="disable")
                Text2 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text2.place(x=408,y=542, height=120)
                Text2.insert(tkinter.INSERT,EEntery5)
                Text2.config(state="disable")
                Text3 =Text(sub_fram , wrap=WORD,width=32,borderwidth=5,bg="#C595C5",foreground= "black",font = ("Arial", 12, "bold"))
                Text3.place(x=705,y=542, height=120)
                Text3.insert(tkinter.INSERT,FEntry5)
                Text3.config(state="disable")     
            except:
                pass

            Well_Lbl= Label(sub_fram,justify="left" ,bg="#5D1C5C" )
            Well_Lbl.place(x=995,y=0,height=1000)
            button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
            text = " Next ",command=lak.Best_Pract,fg= "white", bg="#5D1C5C",width=82,font = ("Arial", 15, "bold"))
            button4.place(x=5,y=755)
            Well_Lbl= Label(sub_fram,justify="left" ,bg="#5D1C5C" )
            Well_Lbl.place(x=0,y=805,width=1100,height=2)
            button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
            text = " Ask Me ",command=lak.main_screen,fg= "white", bg="#5D1C5C",width=26,font = ("Arial", 12, "bold"))
            button4.place(x=723,y=810)
            Well_Lbl= Label(lak.root,justify="left",bg="#69716D"  )
            Well_Lbl.place(x=0,y=805,width=1100,height=2)
                
        Well_Lbl= Label(sub_fram,justify="left" ,bg="#5D1C5C" )
        Well_Lbl.place(x=995,y=0,height=1000)
        button4= Button(lak.root,borderwidth=4,activebackground= "#7FB49A",activeforeground= "white",
        text = " Next ",command=Best_Pract_Element,fg= "white", bg="#5D1C5C",width=82,font = ("Arial", 15, "bold"))
        button4.place(x=5,y=755)
        
        #------------------------------------------------------------------------------------END--------------------------------------------------------------------------------------------------------

    def YourQuestion(lak):
        user= socket.gethostname()
        lak.UWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\User_DB.xlsx")
        S_UserDetail=lak.UWB["User's_detail"]
        VM_no= S_UserDetail["C"]
        VM_list=[]
        for i in VM_no:
            VM_list.append(i.value)
        User= S_UserDetail["A"]
        User_list=[]
        for i in User:
            User_list.append(i.value)
        Team= S_UserDetail["D"]
        Team_list=[]
        for i in Team:
            Team_list.append(i.value)
        VM_ind=VM_list.index(user)
        lak.user= User_list[VM_ind]
        lak.Team= Team_list[VM_ind]

        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")     
        S_Question=lak.QWB["Questions"]
        Csr_name=S_Question["C"]
        Csr_list=[]
        you_name_ind=[]
        your_que_list=[]
        you_date_list=[]
        you_solu_list=[]
        you_Status_list=[]
        for i in Csr_name:
            Csr_list.append(i.value)
        x=len(Csr_list)
        if x==1:
            Well_Lbl= Label(lak.root ,text= "No Records Found",font = ("Arial",20, "bold"),width=15, fg= "black")
            Well_Lbl.place(x=370,y=420)
        else:
            for i in range(x):
                if Csr_list[i]==lak.user:
                    you_name_ind.append(i)
            your_question=S_Question["F"]
            for i in your_question:
                your_que_list.append(i.value)
            you_date=S_Question["A"]
            for i in you_date:
                you_date_list.append(i.value)
            you_Status=S_Question["G"]
            for i in you_Status:
                you_Status_list.append(i.value)
            you_solu=S_Question["H"]
            for i in you_solu:
                you_solu_list.append(i.value)
            
            last_ind= you_name_ind[:-5:-1]

            if len(last_ind)<=1:
                a=last_ind[0]
                ques1= your_que_list[a]
                date1= you_date_list[a]
                status1= you_Status_list[a]
                solution1=you_solu_list[a]
                Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",text= date1,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=5,y=350,height=220)
                Well_Lbl= Label(lak.root, relief= "sunken",bg="white",text= status1,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=160,y=350,height=220)
                Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=41, fg= "black")
                Well_Lbl.place(x=316,y=355,height=216)
                Well_Lbl.insert(tkinter.INSERT,"\n"+ques1)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=41, fg= "green")
                Well_Lbl.place(x=656,y=352,height=216)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution1)
                Well_Lbl.config(state="disabled")
            else:
                a=last_ind[0]
                b=last_ind[1]
                ques1= your_que_list[a]
                ques2= your_que_list[b]
                date1= you_date_list[a]
                date2= you_date_list[b]
                status1= you_Status_list[a]
                status2= you_Status_list[b]
                solution1=you_solu_list[a]
                solution2=you_solu_list[b]
                Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",text= date1,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=5,y=350,height=220)
                Well_Lbl= Label(lak.root, relief= "sunken",bg="white",text= status1,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=160,y=350,height=220)
                Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=41, fg= "black")
                Well_Lbl.place(x=316,y=355,height=216)
                Well_Lbl.insert(tkinter.INSERT,"\n"+ques1)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=41, fg= "green")
                Well_Lbl.place(x=656,y=352,height=216)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution1)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",text= date2,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=5,y=570,height=225)
                Well_Lbl= Label(lak.root, relief= "sunken",bg="white",text= status2,font = ("Arial",12, "bold"),width=15, fg= "black")
                Well_Lbl.place(x=160,y=570,height=225)
                Well_Lbl= Text(lak.root, wrap=WORD, relief= "sunken",borderwidth=5,font = ("Arial", 11, "bold"),width=41, fg= "black")
                Well_Lbl.place(x=316,y=572,height=222)
                Well_Lbl.insert(tkinter.INSERT,"\n"+ques2)
                Well_Lbl.config(state="disabled")
                Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=41, fg= "green")
                Well_Lbl.place(x=656,y=572,height=222)
                Well_Lbl.insert(tkinter.INSERT,"\n"+solution2)
                Well_Lbl.config(state="disabled")
                
                def quenext():
                    if len(last_ind)<=3:
                        a=last_ind[2]
                        ques1= your_que_list[a]
                        date1= you_date_list[a]
                        status1= you_Status_list[a]
                        solution1=you_solu_list[a]
                        Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",text= date1,font = ("Arial",12, "bold"),width=15, fg= "black")
                        Well_Lbl.place(x=5,y=350,height=220)
                        Well_Lbl= Label(lak.root, relief= "sunken",bg="white",text= status1,font = ("Arial",12, "bold"),width=15, fg= "black")
                        Well_Lbl.place(x=160,y=350,height=220)
                        Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=41, fg= "black")
                        Well_Lbl.place(x=316,y=355,height=216)
                        Well_Lbl.insert(tkinter.INSERT,"\n"+ques1)
                        Well_Lbl.config(state="disabled")
                        Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=41, fg= "green")
                        Well_Lbl.place(x=656,y=352,height=216)
                        Well_Lbl.insert(tkinter.INSERT,"\n"+solution1)
                        Well_Lbl.config(state="disabled")
                    else:
                        a=last_ind[2]
                        b=last_ind[3]
                        ques1= your_que_list[a]
                        ques2= your_que_list[b]
                        date1= you_date_list[a]
                        date2= you_date_list[b]
                        status1= you_Status_list[a]
                        status2= you_Status_list[b]
                        solution1=you_solu_list[a]
                        solution2=you_solu_list[b]
                        Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",text= date1,font = ("Arial",12, "bold"),width=15, fg= "black")
                        Well_Lbl.place(x=5,y=350,height=220)
                        Well_Lbl= Label(lak.root, relief= "sunken",bg="white",text= status1,font = ("Arial",12, "bold"),width=15, fg= "black")
                        Well_Lbl.place(x=160,y=350,height=220)
                        Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=41, fg= "black")
                        Well_Lbl.place(x=316,y=355,height=216)
                        Well_Lbl.insert(tkinter.INSERT,"\n"+ques1)
                        Well_Lbl.config(state="disabled")
                        Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken", borderwidth=5,font = ("Arial", 11, "bold"),width=41, fg= "green")
                        Well_Lbl.place(x=656,y=352,height=216)
                        Well_Lbl.insert(tkinter.INSERT,"\n"+solution1)
                        Well_Lbl.config(state="disabled")
                        Well_Lbl= Label(lak.root, relief="sunken" ,bg="white",text= date2,font = ("Arial",12, "bold"),width=15, fg= "black")
                        Well_Lbl.place(x=5,y=570,height=225)
                        Well_Lbl= Label(lak.root, relief= "sunken",bg="white",text= status2,font = ("Arial",12, "bold"),width=15, fg= "black")
                        Well_Lbl.place(x=160,y=570,height=225)
                        Well_Lbl= Text(lak.root, wrap=WORD, relief= "sunken",borderwidth=5,font = ("Arial", 11, "bold"),width=41, fg= "black")
                        Well_Lbl.place(x=316,y=572,height=222)
                        Well_Lbl.insert(tkinter.INSERT,"\n"+ques2)
                        Well_Lbl.config(state="disabled")
                        Well_Lbl= Text(lak.root,wrap=WORD, relief= "sunken",borderwidth=5, font = ("Arial", 11, "bold"),width=41, fg= "green")
                        Well_Lbl.place(x=656,y=572,height=222)
                        Well_Lbl.insert(tkinter.INSERT,"\n"+solution2)
                        Well_Lbl.config(state="disabled")
                    Well_Lbl= Button(lak.root,command=lak.YourQuestion ,text= "Back:",borderwidth=2,justify="left", font = ("Arial",11,"bold"), fg= "black")
                    Well_Lbl.place(x=945,y=286)
                    Well_Lbl= Label(lak.root,bg="#56635C" )
                    Well_Lbl.place(x=0,y=350,width=1000,height=2)
                    Well_Lbl= Label(lak.root,bg="#56635C" )
                    Well_Lbl.place(x=0,y=570,width=1000,height=2)
                    Well_Lbl= Label(lak.root,justify="left" ,bg="#56635C" )
                    Well_Lbl.place(x=995,y=0,height=1000)
                Well_Lbl= Button(lak.root,command=quenext ,text= "Next:",borderwidth=2,justify="left", font = ("Arial",11,"bold"), fg= "black")
                Well_Lbl.place(x=945,y=286)
            Well_Lbl= Label(lak.root,bg="#56635C" )
            Well_Lbl.place(x=0,y=350,width=1000,height=2)
            Well_Lbl= Label(lak.root,bg="#56635C" )
            Well_Lbl.place(x=0,y=570,width=1000,height=2)
            Well_Lbl= Label(lak.root,justify="left" ,bg="#56635C" )
            Well_Lbl.place(x=995,y=0,height=1000)


    def Submit(lak):
        Text_Box=lak.Text_box.get(1.0, END)

        if lak.Team=="Payments & Taxation":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB=op.load_workbook("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")     
        S_Questions=lak.QWB["Questions"]
        #Date	Raised Time	CSR Name	Role	WP	Breif Description	Status	SME-Comments	Close Time	SME (Name)
        x= datetime.datetime.now()
        Date=x.strftime("%d/%m/%Y")
        hour=str(x.strftime("%H"))
        min=str(x.minute)
        Raised_time= hour+":"+min
        Final_list=[Date, Raised_time, lak.user,lak.role.get(),lak.WP.get(),Text_Box,"Pending","Waiting for response","XX:XX",lak.SME.get(),"Date"]
        S_Questions.append(Final_list)
        lak.QWB.save("C:\\Users\\j683\\Desktop\\Question_DB.xlsx")
        if lak.Team=="Payments & Taxation":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\PaymentsQuestion_DB.xlsx")
        elif lak.Team=="Policy Titles":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\TitleQuestion_DB.xlsx")
        elif lak.Team=="Ins & Plan change & Illustration":
            lak.QWB.save("Y:\\ISS\\BWMS\\Q&N\\AdminQuestion_DB.xlsx")     
        lak.Text_box.delete(1.0, END)
        lak.SME.set("")
        lak.role.set("")
        lak.WP.set("")
        lak.YourQuestion()

if __name__== "__main__":
    root=Tk()
    aap= Notification(root)
    aap.start()
    root.mainloop()